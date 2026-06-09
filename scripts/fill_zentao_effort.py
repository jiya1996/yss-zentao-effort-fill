"""
赢时胜禅道工时自包含填报脚本。

本脚本为独立实现。它读取腾讯文档/Excel 工作日志，
按“任务名称”推理禅道项目，逐日填写工时确认页，并在保存后回读校验。
"""

from __future__ import annotations

import argparse
import csv
import io
import json
import os
import re
import sys
from contextlib import contextmanager
from dataclasses import dataclass, field
from datetime import date as Date, datetime
from pathlib import Path
from typing import Any, Iterable, Iterator, Optional


ZENTAO_HOST = os.environ.get("YSS_ZENTAO_HOST", "zentao.example.com")
ZENTAO_URL_PREFIX = f"https://{ZENTAO_HOST.split(':')[0]}"
ZT_EFFORT_URL = (
    f"https://{ZENTAO_HOST}/index.php?m=todo&f=confirmuserconsumed"
    "&day={date}"
)

DEFAULT_SYSTEM = os.environ.get("YSS_ZENTAO_DEFAULT_SYSTEM", "DEFAULT_SYSTEM")
DEFAULT_AI_PROJECT = os.environ.get("YSS_ZENTAO_DEFAULT_PROJECT", "DEFAULT_PROJECT")
DEFAULT_AI_QUERY = os.environ.get("YSS_ZENTAO_DEFAULT_PROJECT_QUERY", DEFAULT_AI_PROJECT)
DEFAULT_WORKLOG_PATH = Path(
    os.environ.get(
        "YSS_WORKLOG_PATH",
        "worklog.xlsx",
    )
)
WORKLOG_SHEET_NAME = "工作日志"
DEFAULT_MAPPING_PATH = Path(__file__).resolve().parents[1] / "工作日志映射.json"


# ---------------------------------------------------------------------------
# 数据模型
# ---------------------------------------------------------------------------

@dataclass
class WorklogEntry:
    date: Date
    start_time: str
    end_time: str
    task_name: str
    content: str

    @property
    def period(self) -> str:
        return infer_period_from_time(self.start_time)

    @property
    def hours(self) -> float:
        return calc_hours_from_range(self.start_time, self.end_time)

    @property
    def is_leave(self) -> bool:
        return bool(re.search(r"请假|休假|年假|病假|调休", self.task_name + self.content))

    @property
    def text(self) -> str:
        return f"{self.task_name} {self.content}".strip()


@dataclass
class ProjectChoice:
    project: str
    query: str
    system: str = DEFAULT_SYSTEM
    source: str = "rule"  # learned | ai_rule | keyword | fallback


@dataclass
class ZentaoEffortGroup:
    date: Date
    project_choice: ProjectChoice
    entries: list[WorklogEntry] = field(default_factory=list)
    name: str = ""
    total_hours: float = 0.0
    system: str = DEFAULT_SYSTEM
    task_type: str = "问题对接"
    description: str = ""


# ---------------------------------------------------------------------------
# 工作日志解析
# ---------------------------------------------------------------------------

FIELD_ALIASES: dict[str, tuple[str, ...]] = {
    "日期": ("日期", "工作日期", "date", "day", "workdate", "work_date"),
    "时间": ("时间", "时段", "起止时间", "time", "timerange", "time_range", "period"),
    "任务名称": ("任务名称", "任务", "名称", "标题", "当日任务", "task", "taskname", "task_name", "name", "title"),
    "工作内容": ("工作内容", "内容", "描述", "说明", "content", "description", "desc", "detail", "details"),
    "当日工时": ("当日工时", "工时", "总工时", "禅道工时", "消耗工时", "实际工时", "hours", "hour"),
}


def _normalize_field_name(name: Any) -> str:
    return re.sub(r"[\s_\-:：/\\]+", "", str(name or "").strip().lower())


def _cell_text(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, (str, int, float)):
        return str(value).strip()
    if isinstance(value, list):
        return " ".join(x for x in (_cell_text(item) for item in value) if x).strip()
    if isinstance(value, dict):
        for key in (
            "text", "plainText", "displayText", "displayValue", "formattedValue",
            "value", "label", "name", "title", "date", "datetime", "timestamp",
            "number", "stringValue", "string_value", "text_value", "number_value",
            "date_value", "datetime_value",
        ):
            if key in value:
                text = _cell_text(value.get(key))
                if text:
                    return text
        for key in ("fieldValue", "field_value", "values", "items", "segments"):
            if key in value:
                text = _cell_text(value.get(key))
                if text:
                    return text
        return ""
    return str(value).strip()


def _extract_record_mapping(record: Any) -> dict[str, Any]:
    if not isinstance(record, dict):
        return {}

    candidates: list[dict[str, Any]] = [record]
    for key in ("fields", "fieldData", "field_data", "values", "data", "record"):
        val = record.get(key)
        if isinstance(val, dict):
            candidates.append(val)

    cells = record.get("cells") or record.get("cellValues") or record.get("cell_values")
    if isinstance(cells, list):
        cell_map: dict[str, Any] = {}
        for cell in cells:
            if not isinstance(cell, dict):
                continue
            key = cell.get("name") or cell.get("title") or cell.get("field") or cell.get("fieldName")
            if key:
                cell_map[str(key)] = cell.get("value", cell.get("text", cell))
        if cell_map:
            candidates.append(cell_map)

    field_values = record.get("field_values") or record.get("fieldValues")
    if isinstance(field_values, list):
        field_map: dict[str, Any] = {}
        for field_value in field_values:
            if not isinstance(field_value, dict):
                continue
            key = (
                field_value.get("field")
                or field_value.get("fieldName")
                or field_value.get("name")
                or field_value.get("title")
            )
            if key:
                field_map[str(key)] = (
                    field_value.get("text_value")
                    or field_value.get("string_value")
                    or field_value.get("number_value")
                    or field_value.get("date_value")
                    or field_value.get("datetime_value")
                    or field_value.get("value")
                    or field_value
                )
        if field_map:
            candidates.append(field_map)

    merged: dict[str, Any] = {}
    ignored = {"id", "recordId", "record_id", "rowId", "row_id", "createdAt", "updatedAt"}
    for cand in candidates:
        for key, value in cand.items():
            if key in ignored:
                continue
            merged[str(key)] = value
    return merged


def _pick_field(record: dict[str, Any], canonical_name: str) -> str:
    aliases = {_normalize_field_name(x) for x in FIELD_ALIASES[canonical_name]}
    for key, value in record.items():
        if _normalize_field_name(key) in aliases:
            return _cell_text(value)
    return ""


def _records_from_json_payload(payload: Any) -> list[Any]:
    if isinstance(payload, list):
        return payload
    if isinstance(payload, dict):
        normalized_keys = {_normalize_field_name(k) for k in payload.keys()}
        if normalized_keys & {_normalize_field_name(x) for x in FIELD_ALIASES["日期"]}:
            return [payload]
        for key in (
            "records", "structuredContent", "structured_content", "result",
            "rows", "data", "list", "items", "values",
        ):
            if key in payload:
                rows = _records_from_json_payload(payload[key])
                if rows:
                    return rows
    return []


def parse_date_cell(value: Any) -> Date:
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, Date):
        return value

    text = str(value or "").strip()
    if not text:
        raise ValueError("日期为空")

    if re.fullmatch(r"\d{10,}", text):
        ts = int(text) / 1000 if len(text) > 10 else int(text)
        return datetime.fromtimestamp(ts).date()

    cn = re.fullmatch(r"(\d{4})年\s*(\d{1,2})月\s*(\d{1,2})日?", text)
    if cn:
        return Date(int(cn.group(1)), int(cn.group(2)), int(cn.group(3)))

    for fmt in ("%Y-%m-%d", "%Y/%m/%d", "%Y.%m.%d", "%m/%d", "%m-%d"):
        try:
            parsed = datetime.strptime(text, fmt)
            if fmt in ("%m/%d", "%m-%d"):
                return Date(Date.today().year, parsed.month, parsed.day)
            return parsed.date()
        except ValueError:
            continue
    raise ValueError(f"无法解析日期：{text!r}")


def _normalize_time_text(text: str) -> str:
    return (
        str(text or "")
        .strip()
        .replace("：", ":")
        .replace("－", "-")
        .replace("—", "-")
        .replace("–", "-")
        .replace("~", "-")
    )


def _format_hhmm(text: str) -> str:
    h, m = text.split(":")
    return f"{int(h):02d}:{int(m):02d}"


def parse_time_range(text: str) -> tuple[str, str]:
    normalized = _normalize_time_text(text)
    match = re.match(r"(\d{1,2}:\d{1,2})\s*-\s*(\d{1,2}:\d{1,2})", normalized)
    if not match:
        raise ValueError(f"无法解析时间区间：{text!r}，格式应为 HH:MM-HH:MM")
    return _format_hhmm(match.group(1)), _format_hhmm(match.group(2))


def _time_range_for_hours(hours: float) -> tuple[str, str]:
    start_minutes = 9 * 60
    end_minutes = start_minutes + int(round(hours * 60))
    end_h, end_m = divmod(end_minutes, 60)
    return "09:00", f"{end_h:02d}:{end_m:02d}"


def _parse_hours(value: str) -> Optional[float]:
    match = re.search(r"\d+(?:\.\d+)?", str(value or ""))
    if not match:
        return None
    hours = float(match.group(0))
    return hours if hours > 0 else None


def calc_hours_from_range(start: str, end: str) -> float:
    sh, sm = int(start[:2]), int(start[3:5])
    eh, em = int(end[:2]), int(end[3:5])
    start_min = sh * 60 + sm
    end_min = eh * 60 + em
    total_min = end_min - start_min
    if total_min <= 0:
        raise ValueError(f"结束时间 {end} 不晚于开始时间 {start}")

    for break_start, break_end in ((12 * 60, 13 * 60), (18 * 60, 19 * 60)):
        if start_min <= break_start and end_min >= break_end:
            total_min -= break_end - break_start
    return int((total_min / 60.0) * 2) / 2


def infer_period_from_time(start_time: str) -> str:
    h, m = int(start_time[:2]), int(start_time[3:5])
    if h < 12:
        return "上午"
    if h < 18 or (h == 18 and m == 0):
        return "下午"
    return "晚上"


def _entry_from_row(row: dict[str, Any]) -> Optional[WorklogEntry]:
    raw_date = _pick_field(row, "日期")
    time_text = _pick_field(row, "时间")
    task_name = _pick_field(row, "任务名称")
    content = _pick_field(row, "工作内容")

    if not any([raw_date, time_text, task_name, content]):
        return None

    day = parse_date_cell(raw_date)
    if time_text:
        start_t, end_t = parse_time_range(time_text)
    else:
        hours = _parse_hours(_pick_field(row, "当日工时"))
        if hours is None:
            return None
        start_t, end_t = _time_range_for_hours(hours)

    effective_task_name = (task_name or content).strip()
    if not effective_task_name:
        return None

    return WorklogEntry(
        date=day,
        start_time=start_t,
        end_time=end_t,
        task_name=effective_task_name,
        content=content.strip(),
    )


def _filter_entries(entries: Iterable[WorklogEntry],
                    start: Optional[Date],
                    end: Optional[Date]) -> list[WorklogEntry]:
    result = []
    for entry in entries:
        if start and entry.date < start:
            continue
        if end and entry.date > end:
            continue
        result.append(entry)
    result.sort(key=lambda e: (e.date, e.start_time))
    return result


def load_worklog_from_records(records: Any,
                              start: Optional[Date] = None,
                              end: Optional[Date] = None) -> list[WorklogEntry]:
    entries: list[WorklogEntry] = []
    for record in _records_from_json_payload(records):
        row = _extract_record_mapping(record)
        try:
            entry = _entry_from_row(row)
        except ValueError:
            continue
        if entry:
            entries.append(entry)
    return _filter_entries(entries, start, end)


def load_worklog_from_json(path: str | Path,
                           start: Optional[Date] = None,
                           end: Optional[Date] = None) -> list[WorklogEntry]:
    if str(path) == "-":
        payload = json.loads(sys.stdin.read())
    else:
        p = Path(path)
        if not p.exists():
            raise FileNotFoundError(f"JSON 工作日志不存在：{p}")
        payload = json.loads(p.read_text(encoding="utf-8-sig"))
    return load_worklog_from_records(payload, start, end)


def load_worklog_from_csv(path: str | Path,
                          start: Optional[Date] = None,
                          end: Optional[Date] = None) -> list[WorklogEntry]:
    if str(path).startswith("csv://"):
        text = str(path)[6:]
    else:
        p = Path(path)
        if not p.exists():
            raise FileNotFoundError(f"CSV 工作日志不存在：{p}")
        text = p.read_text(encoding="utf-8-sig")
    reader = csv.DictReader(io.StringIO(text))
    return load_worklog_from_records(list(reader), start, end)


def load_worklog(path: str | Path | None,
                 start: Optional[Date] = None,
                 end: Optional[Date] = None) -> list[WorklogEntry]:
    try:
        import openpyxl
    except ImportError as exc:
        raise RuntimeError("读取 Excel 需要安装 openpyxl") from exc

    xlsx = Path(path) if path else DEFAULT_WORKLOG_PATH
    if not xlsx.exists():
        raise FileNotFoundError(f"工作日志不存在：{xlsx}")

    wb = openpyxl.load_workbook(xlsx, data_only=True)
    if WORKLOG_SHEET_NAME not in wb.sheetnames:
        raise ValueError(f"工作日志缺少 sheet：{WORKLOG_SHEET_NAME}")
    ws = wb[WORKLOG_SHEET_NAME]

    headers: dict[str, int] = {}
    for idx, cell in enumerate(ws[1], start=1):
        name = str(cell.value or "").strip()
        if name:
            headers[name] = idx

    rows: list[dict[str, Any]] = []
    for row_idx in range(2, ws.max_row + 1):
        row: dict[str, Any] = {}
        for name, col in headers.items():
            row[name] = ws.cell(row_idx, col).value
        rows.append(row)
    return load_worklog_from_records(rows, start, end)


# ---------------------------------------------------------------------------
# 项目推理与聚合
# ---------------------------------------------------------------------------

class MappingStore:
    def __init__(self, path: str | Path | None = None):
        self.path = Path(path) if path else DEFAULT_MAPPING_PATH
        self.data = self._load()

    def _load(self) -> dict[str, Any]:
        if not self.path.exists():
            return {"task_to_project": {}}
        try:
            return json.loads(self.path.read_text(encoding="utf-8"))
        except json.JSONDecodeError:
            return {"task_to_project": {}}

    def save(self) -> None:
        self.path.parent.mkdir(parents=True, exist_ok=True)
        self.path.write_text(json.dumps(self.data, ensure_ascii=False, indent=2), encoding="utf-8")

    @staticmethod
    def _key(task_name: str) -> str:
        return re.sub(r"\s+", "", task_name.strip().lower())

    def get_project_by_task(self, task_name: str) -> Optional[ProjectChoice]:
        key = self._key(task_name)
        if not key:
            return None
        records = self.data.get("task_to_project", {})
        rec = records.get(key)
        if not rec:
            for learned_key, value in records.items():
                if key.startswith(learned_key) or learned_key.startswith(key):
                    rec = value
                    break
        if not rec:
            return None
        return ProjectChoice(
            project=rec.get("project", ""),
            query=rec.get("query", ""),
            system=rec.get("system", DEFAULT_SYSTEM),
            source="learned",
        )

    def learn_task_project(self, task_name: str, project: str, query: str,
                           system: str = DEFAULT_SYSTEM) -> None:
        key = self._key(task_name)
        if not key:
            return
        self.data.setdefault("task_to_project", {})[key] = {
            "project": project,
            "query": query,
            "system": system,
        }
        self.save()


AI_PROJECT_PATTERN = re.compile(
    r"AI|ai|知识库|数字员工|wiki|knowledge|platform|automation|agent|大模型|智能体|本体|赋能|提效",
    re.IGNORECASE,
)
CUSTOMER_PATTERN = re.compile(
    r"([\u4e00-\u9fa5A-Za-z0-9]{2,20}(?:基金|银行|证券|保险|信托|资管|期货|客户|项目))"
)


def infer_project(entry: WorklogEntry, mapping: Optional[MappingStore] = None) -> ProjectChoice:
    task_text = entry.task_name.strip()

    if mapping:
        learned = mapping.get_project_by_task(task_text)
        if learned:
            return learned

    if AI_PROJECT_PATTERN.search(task_text):
        return ProjectChoice(DEFAULT_AI_PROJECT, DEFAULT_AI_QUERY, DEFAULT_SYSTEM, "ai_rule")

    customer_match = CUSTOMER_PATTERN.search(task_text)
    if customer_match:
        query = customer_match.group(1)
        return ProjectChoice("", query, DEFAULT_SYSTEM, "keyword")

    return ProjectChoice(DEFAULT_AI_PROJECT, DEFAULT_AI_QUERY, DEFAULT_SYSTEM, "fallback")


TASK_TYPE_RULES = [
    (r"文档|方案|材料|demo|演示|整理|撰写|编写|PPT", "文档编写"),
    (r"客户|回访|沟通|交流|汇报", "客户回访"),
    (r"问题|运维|排查|修复|定位|支持|对接", "问题对接"),
    (r"需求|分析|评审|梳理|复盘|调研|设计", "需求分析"),
    (r"测试|验证|联调|自测|评测", "测试验证"),
    (r"请假|休假|年假|病假|调休", "其他"),
]


def infer_task_type_from_text(text: str) -> str:
    for pattern, task_type in TASK_TYPE_RULES:
        if re.search(pattern, text, re.IGNORECASE):
            return task_type
    return "问题对接"


def group_entries_by_date(entries: Iterable[WorklogEntry]) -> dict[Date, list[WorklogEntry]]:
    result: dict[Date, list[WorklogEntry]] = {}
    for entry in entries:
        result.setdefault(entry.date, []).append(entry)
    return result


def build_zentao_effort_groups(entries: list[WorklogEntry],
                               mapping: Optional[MappingStore]) -> list[ZentaoEffortGroup]:
    buckets: dict[tuple[Date, str, str], tuple[ProjectChoice, list[WorklogEntry]]] = {}

    for entry in entries:
        if entry.is_leave:
            continue
        choice = infer_project(entry, mapping)
        key = choice.project or choice.query or DEFAULT_AI_PROJECT
        bucket_key = (entry.date, key, choice.source)
        if bucket_key not in buckets:
            buckets[bucket_key] = (choice, [])
        buckets[bucket_key][1].append(entry)

    groups: list[ZentaoEffortGroup] = []
    for (day, key, _source), (choice, bucket) in sorted(buckets.items(), key=lambda x: (x[0][0], x[0][1])):
        bucket.sort(key=lambda e: e.start_time)
        names = list(dict.fromkeys(e.task_name for e in bucket))
        name = names[0] if len(names) == 1 else "、".join(names)
        total_hours = sum(e.hours for e in bucket)
        task_type = infer_task_type_from_text(" ".join(e.task_name for e in bucket))

        period_parts: dict[str, list[str]] = {}
        for entry in bucket:
            text = (entry.content or entry.task_name).strip()
            if text:
                period_parts.setdefault(entry.period, []).append(text)

        desc_bits = []
        for period in ("上午", "下午", "晚上"):
            if period in period_parts:
                desc_bits.append(f"{period}：{'、'.join(dict.fromkeys(period_parts[period]))}")
        description = "；".join(desc_bits)

        groups.append(ZentaoEffortGroup(
            date=day,
            project_choice=choice,
            entries=bucket,
            name=name,
            total_hours=total_hours,
            system=choice.system,
            task_type=task_type,
            description=description,
        ))

    return groups


def _calc_leave_hours_for_date(entries: list[WorklogEntry]) -> float:
    return sum(e.hours for e in entries if e.is_leave)


# ---------------------------------------------------------------------------
# 浏览器与页面操作
# ---------------------------------------------------------------------------

@contextmanager
def get_browser_context() -> Iterator[Any]:
    try:
        from playwright.sync_api import sync_playwright
    except ImportError as exc:
        raise RuntimeError("缺少 playwright，请先安装 playwright 并启动 ChromeDebug") from exc

    with sync_playwright() as pw:
        browser = pw.chromium.connect_over_cdp("http://127.0.0.1:9222")
        try:
            context = browser.contexts[0] if browser.contexts else browser.new_context()
            yield context
        finally:
            browser.close()


def find_or_open_tab(context: Any, url_prefix: str, target_url: str) -> Any:
    for page in context.pages:
        try:
            if page.url.startswith(url_prefix):
                return page
        except Exception:
            continue
    return context.new_page()


def _read_existing_rows(page: Any) -> list[dict[str, Any]]:
    raw = page.evaluate(
        """() => JSON.stringify(Array.from(
          document.querySelectorAll('table.table-1 tbody tr, table.colored tbody tr')
        ).map((tr, index) => {
          const tds = Array.from(tr.querySelectorAll('td'));
          function val(idx) {
            const td = tds[idx];
            if (!td) return '';
            const input = td.querySelector('input[type=text], textarea');
            if (input) return (input.value || '').trim();
            const select = td.querySelector('select');
            if (select && select.selectedOptions && select.selectedOptions[0]) {
              return select.selectedOptions[0].text.trim();
            }
            return (td.innerText || td.textContent || '').trim().split('\\n').filter(Boolean)[0] || '';
          }
          return {
            index,
            col_count: tds.length,
            name: val(3),
            hours: parseFloat(val(7)) || 0,
            project: val(8),
            system: val(10),
            task_type: val(11),
            description: val(12),
            all_text: (tr.innerText || tr.textContent || '').trim()
          };
        }).filter(row => row.col_count >= 13))"""
    )
    return json.loads(raw) if raw else []


def _find_row_index(existing: list[dict[str, Any]], name: str, used: set[int]) -> Optional[int]:
    for row in existing:
        if row["index"] in used:
            continue
        if row.get("name") == name:
            return int(row["index"])
    for row in existing:
        if row["index"] in used:
            continue
        if not row.get("name") or float(row.get("hours") or 0) == 0:
            return int(row["index"])
    return None


def _click_add_row(page: Any, row_index: Optional[int] = None) -> None:
    clicked = page.evaluate(
        """(rowIndex) => {
          const rows = Array.from(
            document.querySelectorAll('table.table-1 tbody tr, table.colored tbody tr')
          ).filter(tr => tr.querySelectorAll('td').length >= 13);
          const tr = rowIndex == null ? rows[rows.length - 1] : rows[rowIndex];
          if (!tr) return false;
          const cells = Array.from(tr.querySelectorAll('td'));
          const rightCells = cells.slice(Math.max(0, cells.length - 2));
          const candidates = Array.from(
            (rightCells.length ? rightCells : cells).flatMap(td => Array.from(td.querySelectorAll('a, button, i, span')))
          ).filter(el => {
            const text = (el.innerText || el.textContent || '').trim();
            const title = (el.getAttribute('title') || el.getAttribute('aria-label') || '').trim();
            const cls = (el.className || '').toString();
            const onclick = (el.getAttribute('onclick') || '').toString();
            const href = (el.getAttribute('href') || '').toString();
            return text === '+'
              || /新增|添加|add/i.test(text + title + cls + onclick + href)
              || /icon-(plus|add)|btn-(add|plus)/i.test(cls);
          });
          const preferred = candidates[0];
          if (!preferred) return false;
          const target = preferred.tagName === 'I' || preferred.tagName === 'SPAN'
            ? (preferred.closest('a, button') || preferred)
            : preferred;
          target.click();
          return true;
        }""",
        row_index,
    )
    if clicked:
        page.wait_for_timeout(900)
        return

    for selector in [
        "a:has-text('+')",
        "button:has-text('+')",
        "a:has-text('新增')",
        "button:has-text('新增')",
        ".icon-plus",
    ]:
        try:
            page.locator(selector).last.click(timeout=2000)
            page.wait_for_timeout(900)
            return
        except Exception:
            continue
    raise RuntimeError("找不到禅道工时新增行按钮")


def _set_text_cell(page: Any, row_index: int, col_index: int, text: str) -> None:
    page.evaluate(
        """([rowIndex, colIndex, text]) => {
          const rows = Array.from(
            document.querySelectorAll('table.table-1 tbody tr, table.colored tbody tr')
          ).filter(tr => tr.querySelectorAll('td').length >= 13);
          const tr = rows[rowIndex];
          if (!tr) throw new Error(`row ${rowIndex} not found`);
          const td = tr.querySelectorAll('td')[colIndex];
          if (!td) throw new Error(`col ${colIndex} not found`);
          let input = td.querySelector('input[type=text], textarea');
          if (!input) {
            input = document.createElement('input');
            input.type = 'text';
            td.appendChild(input);
          }
          input.value = text;
          input.dispatchEvent(new Event('input', {bubbles:true}));
          input.dispatchEvent(new Event('change', {bubbles:true}));
        }""",
        [row_index, col_index, text],
    )


def _row_cell(page: Any, row_index: int, col_index: int) -> Any:
    return (
        page.locator("table.table-1 tbody tr, table.colored tbody tr")
        .nth(row_index)
        .locator("td")
        .nth(col_index)
    )


def _cell_input_value(page: Any, row_index: int, col_index: int) -> str:
    try:
        return page.evaluate(
            """([rowIndex, colIndex]) => {
              const rows = Array.from(
                document.querySelectorAll('table.table-1 tbody tr, table.colored tbody tr')
              ).filter(tr => tr.querySelectorAll('td').length >= 13);
              const tr = rows[rowIndex];
              if (!tr) return '';
              const td = tr.querySelectorAll('td')[colIndex];
              if (!td) return '';
              const input = td.querySelector('input[type=text], textarea');
              if (input) return (input.value || '').trim();
              const select = td.querySelector('select');
              if (select && select.selectedOptions && select.selectedOptions[0]) {
                return select.selectedOptions[0].text.trim();
              }
              return (td.innerText || td.textContent || '').trim();
            }""",
            [row_index, col_index],
        )
    except Exception:
        return ""


def _is_blank_cell_value(value: str) -> bool:
    text = re.sub(r"\s+", "", value or "")
    return text in {"", "-", "--", "请选择", "无"}


def _compact_note_text(text: str, limit: int = 600) -> str:
    lines = [line.strip() for line in (text or "").splitlines() if line.strip()]
    compact = "；".join(lines)
    if len(compact) > limit:
        return compact[:limit] + "..."
    return compact


def _close_maintenance_note(page: Any) -> None:
    for selector in [
        "[class*='modal'] button:has-text('关闭')",
        "[class*='modal'] button:has-text('确定')",
        "[class*='modal'] .close",
        "[role=dialog] button:has-text('关闭')",
        "[role=dialog] button:has-text('确定')",
        ".layui-layer-close",
    ]:
        try:
            loc = page.locator(selector)
            if loc.count() > 0:
                loc.first.click(timeout=1000)
                page.wait_for_timeout(300)
                return
        except Exception:
            continue
    try:
        page.keyboard.press("Escape")
        page.wait_for_timeout(300)
    except Exception:
        pass


def _read_maintenance_note(page: Any, field_label: str) -> str:
    """点击表头红色“维护说明”，读取项目/对应系统的公司维护规则。"""
    try:
        clicked = page.evaluate(
            """(fieldLabel) => {
              const textOf = el => (el.innerText || el.textContent || '').trim();
              const cells = Array.from(document.querySelectorAll('th, td')).filter(el => {
                const text = textOf(el).replace(/\\s+/g, '');
                return text.includes(fieldLabel) && text.includes('维护说明');
              });
              const scope = cells[0];
              if (!scope) return false;

              const targets = Array.from(scope.querySelectorAll('a, button, span, font, em, b, i, div'))
                .filter(el => textOf(el).replace(/\\s+/g, '').includes('维护说明'));
              const target = targets.find(el => {
                const style = window.getComputedStyle(el);
                const onclick = el.getAttribute('onclick') || '';
                const role = el.getAttribute('role') || '';
                return onclick || role === 'button' || style.cursor === 'pointer' || el.closest('a, button');
              }) || targets[0] || scope;

              const clickable = target.closest('a, button') || target;
              clickable.scrollIntoView({block: 'center', inline: 'center'});
              clickable.click();
              return true;
            }""",
            field_label,
        )
        if not clicked:
            return ""
        page.wait_for_timeout(900)
        note = page.evaluate(
            """() => {
              const isVisible = el => {
                const style = window.getComputedStyle(el);
                const rect = el.getBoundingClientRect();
                return style.display !== 'none'
                  && style.visibility !== 'hidden'
                  && rect.width > 0
                  && rect.height > 0;
              };
              const selectors = [
                '[role=dialog]',
                '.modal',
                '.bootbox',
                '.popover',
                '.layui-layer',
                '.messager-window',
                '.dialog'
              ];
              const candidates = [];
              for (const selector of selectors) {
                candidates.push(...Array.from(document.querySelectorAll(selector)));
              }
              const texts = candidates
                .filter(isVisible)
                .map(el => (el.innerText || el.textContent || '').trim())
                .filter(text => text.length > 8);
              texts.sort((a, b) => b.length - a.length);
              return texts[0] || '';
            }"""
        )
        _close_maintenance_note(page)
        return _compact_note_text(note)
    except Exception as exc:
        print(f"    ⚠️ 读取{field_label}维护说明失败：{exc}")
        return ""


def _log_maintenance_note(page: Any, field_label: str, reason: str) -> None:
    note = _read_maintenance_note(page, field_label)
    if note:
        print(f"    ℹ️ 已参考{field_label}维护说明（{reason}）：{note}")
    else:
        print(f"    ⚠️ 未能读取{field_label}维护说明（{reason}）")


def _select_dropdown_cell(page: Any, row_index: int, col_index: int,
                          query: str, prefer_text: str = "",
                          optional: bool = False) -> str:
    if not query and not prefer_text:
        return ""

    cell = _row_cell(page, row_index, col_index)

    # 优先尝试原生 <select> 元素（含重试：等待 AJAX 加载选项）
    for attempt in range(8):
        native_result = page.evaluate(
            """([rowIndex, colIndex, target, preferText]) => {
              const rows = Array.from(
                document.querySelectorAll('table.table-1 tbody tr, table.colored tbody tr')
              ).filter(tr => tr.querySelectorAll('td').length >= 13);
              const tr = rows[rowIndex];
              if (!tr) return {found: false, reason: 'no_row'};
              const td = tr.querySelectorAll('td')[colIndex];
              if (!td) return {found: false, reason: 'no_td'};
              const select = td.querySelector('select');
              if (!select) return {found: false, reason: 'no_select'};
              const options = Array.from(select.options);
              const optionTexts = options.map(o => o.text.trim()).filter(Boolean);
              if (optionTexts.length === 0) return {found: true, matched: null, options: [], empty: true};
              const matchTexts = preferText ? [preferText, target] : [target];
              for (const mt of matchTexts) {
                const idx = options.findIndex(o => o.text.trim().toLowerCase().includes(mt.toLowerCase())
                  || mt.toLowerCase().includes(o.text.trim().toLowerCase()));
                if (idx >= 0) {
                  select.selectedIndex = idx;
                  select.dispatchEvent(new Event('change', {bubbles:true}));
                  return {found: true, matched: options[idx].text.trim()};
                }
              }
              return {found: true, matched: null, options: optionTexts.slice(0, 30)};
            }""",
            [row_index, col_index, query, prefer_text],
        )
        if not isinstance(native_result, dict):
            break
        if not native_result.get("found"):
            break
        if native_result.get("matched"):
            page.wait_for_timeout(300)
            return native_result["matched"]
        if native_result.get("empty"):
            # 选项尚未加载，等待 AJAX 后重试
            if attempt < 7:
                page.wait_for_timeout(800)
                continue
            if optional:
                return ""
            print(f"    ⚠️ 下拉选项始终为空 col={col_index}, 已等待 {attempt+1} 次")
            break
        # 有选项但未匹配
        print(f"    ⚠️ 下拉选项列表: {native_result.get('options', [])}, 未匹配: query={query!r}")
        break

    try:
        existing = cell.inner_text(timeout=1000).strip()
        if existing and existing != "-":
            return existing
    except Exception:
        pass

    cell.click(timeout=5000)
    page.wait_for_timeout(600)

    target = prefer_text or query
    for selector in [
        "input.select2-input:visible",
        ".select2-drop-active input:visible",
        ".select2-search input:visible",
        ".chosen-search input:visible",
        "input[type=text]:visible",
    ]:
        try:
            loc = page.locator(selector)
            if loc.count() == 0:
                continue
            loc.last.fill(target, timeout=2000)
            page.wait_for_timeout(1500)
            break
        except Exception:
            continue

    clicked = False
    option_selectors = [".select2-result-label", ".select2-results li", ".chosen-results li"]
    for match_text in ([prefer_text] if prefer_text else []) + ([query] if query else []):
        for selector in option_selectors:
            try:
                options = page.locator(selector)
                for idx in range(min(options.count(), 20)):
                    text = options.nth(idx).inner_text(timeout=1000)
                    if match_text.lower() in text.lower():
                        options.nth(idx).click(timeout=3000)
                        clicked = True
                        break
                if clicked:
                    break
            except Exception:
                continue
        if clicked:
            break

    if not clicked:
        if optional:
            return _cell_input_value(page, row_index, col_index)
        page.keyboard.press("Enter")
    page.wait_for_timeout(800)

    try:
        return cell.inner_text(timeout=2000).strip() or target
    except Exception:
        return target


def _trigger_project_change(page: Any, row_index: int) -> None:
    page.evaluate(
        """(rowIndex) => {
          const rows = Array.from(
            document.querySelectorAll('table.table-1 tbody tr, table.colored tbody tr')
          ).filter(tr => tr.querySelectorAll('td').length >= 13);
          const tr = rows[rowIndex];
          if (!tr) return;
          const td = tr.querySelectorAll('td')[8];
          if (!td) return;
          const input = td.querySelector('input[type=hidden], input[name^="projects"]');
          if (!input) return;
          input.dispatchEvent(new Event('input', {bubbles:true}));
          input.dispatchEvent(new Event('change', {bubbles:true}));
          if (window.jQuery) {
            try {
              window.jQuery(input).trigger('change');
              window.jQuery(input).trigger('select2-selected');
            } catch (e) {}
          }
        }""",
        row_index,
    )
    # 等待系统列自动带出（禅道 AJAX 刷新）
    page.wait_for_timeout(1500)


def _wait_for_auto_system(page: Any, row_index: int, timeout_ms: int = 5000) -> str:
    """等待项目选择后禅道自动带出“对应系统”。

    禅道在选项目后会异步刷新系统列。系统列不是本脚本的主控字段：
    能带出就回读校验，带不出时继续保存前的其他校验，避免卡死。
    """
    step_ms = 500
    attempts = max(1, timeout_ms // step_ms)
    for _ in range(attempts):
        value = _cell_input_value(page, row_index, 10)
        if not _is_blank_cell_value(value):
            return value
        page.wait_for_timeout(step_ms)
    return _cell_input_value(page, row_index, 10)


def _ensure_system_value(page: Any, row_index: int, preferred_system: str) -> str:
    auto_system = _wait_for_auto_system(page, row_index)
    if not _is_blank_cell_value(auto_system):
        print(f"    ℹ️ 对应系统自动带出：{auto_system}")
        return auto_system

    fallback_system = preferred_system or DEFAULT_SYSTEM
    _log_maintenance_note(page, "对应系统", "项目未自动带出对应系统")
    print(f"    ⚠️ 项目未自动带出对应系统，改用兜底系统：{fallback_system}")
    selected = _select_dropdown_cell(page, row_index, 10, fallback_system, fallback_system)
    final_value = _cell_input_value(page, row_index, 10) or selected
    if _is_blank_cell_value(final_value):
        raise RuntimeError(f"对应系统未自动带出，且无法选择兜底系统：{fallback_system}")
    return final_value


def _project_matches(actual: str, expected: str) -> bool:
    actual_norm = re.sub(r"\s+", "", actual or "")
    expected_norm = re.sub(r"\s+", "", expected or "")
    return bool(
        actual_norm == expected_norm
        or expected_norm in actual_norm
        or actual_norm in expected_norm
    )


def _fill_effort_row_on_page(page: Any, row_index: int, group: ZentaoEffortGroup) -> str:
    _set_text_cell(page, row_index, 3, group.name)
    _set_text_cell(page, row_index, 7, f"{group.total_hours:g}")
    selected_project = _select_dropdown_cell(
        page,
        row_index,
        8,
        group.project_choice.query,
        group.project_choice.project,
    )
    if group.project_choice.project and not _project_matches(selected_project, group.project_choice.project):
        _log_maintenance_note(page, "项目", "项目选择不匹配")
        raise RuntimeError(
            f"项目选择不匹配：实际={selected_project!r}，期望={group.project_choice.project!r}"
        )
    _trigger_project_change(page, row_index)
    _ensure_system_value(page, row_index, group.system)
    selected_task_type = _select_dropdown_cell(
        page, row_index, 11, group.task_type, group.task_type, optional=True
    )
    if _is_blank_cell_value(selected_task_type):
        print("    ⚠️ 任务类型选项未加载；跳过任务类型手工选择")
    if group.description:
        _set_text_cell(page, row_index, 12, group.description)
    return selected_project


def _validate_filled_row(page: Any, row_index: int, group: ZentaoEffortGroup,
                         selected_project: str) -> None:
    checks = [
        ("名称", 3, group.name),
        ("工时", 7, f"{group.total_hours:g}"),
    ]
    if group.description:
        checks.append(("描述", 12, group.description))

    errors = []
    project_actual = _cell_input_value(page, row_index, 8) or selected_project
    project_expected = group.project_choice.project or selected_project or group.project_choice.query
    if project_expected and not _project_matches(project_actual, project_expected):
        errors.append(f"项目={project_actual!r} != {project_expected!r}")

    system_actual = _cell_input_value(page, row_index, 10)
    if _is_blank_cell_value(system_actual):
        print(f"    ⚠️ 对应系统为空：{group.name}。已继续校验其他字段")

    task_type_actual = _cell_input_value(page, row_index, 11)
    if _is_blank_cell_value(task_type_actual):
        print(f"    ⚠️ 任务类型为空：{group.name}。已继续校验其他字段")
    elif (task_type_actual or "").strip() != (group.task_type or "").strip():
        errors.append(f"任务类型={task_type_actual!r} != {group.task_type!r}")

    for label, col, expected in checks:
        actual = _cell_input_value(page, row_index, col)
        if label == "工时":
            try:
                if abs(float(actual or 0) - float(expected or 0)) > 0.01:
                    errors.append(f"{label}={actual!r} != {expected!r}")
            except Exception:
                errors.append(f"{label}={actual!r} != {expected!r}")
        elif (actual or "").strip() != (expected or "").strip():
            errors.append(f"{label}={actual!r} != {expected!r}")

    if errors:
        raise RuntimeError("禅道工时填写校验失败: " + "；".join(errors))


def _save_page_and_confirm(page: Any) -> None:
    page.evaluate("() => { window.confirm = () => true; window.alert = () => {}; }")
    page.on("dialog", lambda dialog: dialog.accept())

    save_clicked = False
    for selector in ["button:has-text('保存')", "input[value='保存']", "a:has-text('保存')"]:
        try:
            page.locator(selector).first.click(timeout=3000)
            save_clicked = True
            break
        except Exception:
            continue
    if not save_clicked:
        raise RuntimeError("找不到保存按钮")

    page.wait_for_timeout(3000)

    # 检测禅道校验错误（如必填字段未填）
    error_msgs = page.evaluate(
        """() => {
          const errors = [];
          // 检查红色提示 / alert-danger / 表单校验错误
          document.querySelectorAll('.text-danger, .form-error, .error, .alert-danger, .tips-error').forEach(el => {
            const t = (el.innerText || el.textContent || '').trim();
            if (t) errors.push(t);
          });
          // 检查 Zentao 特有的提示框
          document.querySelectorAll('.modal.in .modal-body, .bootbox .modal-body').forEach(el => {
            const t = (el.innerText || el.textContent || '').trim();
            if (t && t.length < 200) errors.push(t);
          });
          return errors;
        }"""
    )
    if error_msgs:
        print(f"    ⚠️ 保存后检测到提示信息: {error_msgs}")

    for selector in [
        "[class*='modal'] button:has-text('确认')",
        "[class*='dialog'] button:has-text('确认')",
        "[class*='modal'] button:has-text('确定')",
        "[class*='dialog'] button:has-text('确定')",
        ".btn-primary:has-text('确认')",
        ".btn-primary:has-text('确定')",
    ]:
        try:
            loc = page.locator(selector)
            if loc.count() > 0:
                loc.first.click(timeout=2000)
                page.wait_for_timeout(1000)
                break
        except Exception:
            continue

    # 保存后二次检测错误（某些错误在确认后才出现）
    page.wait_for_timeout(2000)
    post_errors = page.evaluate(
        """() => {
          const errors = [];
          document.querySelectorAll('.text-danger, .form-error, .error, .alert-danger, .tips-error').forEach(el => {
            const t = (el.innerText || el.textContent || '').trim();
            if (t) errors.push(t);
          });
          return errors;
        }"""
    )
    if post_errors:
        raise RuntimeError("保存后检测到错误: " + "；".join(post_errors))


def _validate_saved_row(page: Any, row_index: int, group: ZentaoEffortGroup,
                        selected_project: str) -> None:
    """保存并刷新后校验持久化结果。"""
    errors = []

    actual_hours = _cell_input_value(page, row_index, 7)
    try:
        if abs(float(actual_hours or 0) - group.total_hours) > 0.01:
            errors.append(f"工时={actual_hours!r} != {group.total_hours:g}")
    except Exception:
        errors.append(f"工时={actual_hours!r} != {group.total_hours:g}")

    project_actual = _cell_input_value(page, row_index, 8)
    project_expected = group.project_choice.project or selected_project or group.project_choice.query
    if project_expected and not _project_matches(project_actual, project_expected):
        errors.append(f"项目={project_actual!r} != {project_expected!r}")

    if errors:
        raise RuntimeError("保存并刷新后回读校验失败: " + "；".join(errors))
    print(f"    ✅ 保存并刷新后回读通过：工时={actual_hours}h, 项目={project_actual}")


def _find_saved_row_index(rows: list[dict[str, Any]],
                          group: ZentaoEffortGroup,
                          selected_project: str,
                          used: set[int]) -> Optional[int]:
    for row in rows:
        idx = int(row["index"])
        if idx in used:
            continue
        if str(row.get("name") or "").strip() == group.name:
            return idx

    project_expected = group.project_choice.project or selected_project or group.project_choice.query
    candidates: list[int] = []
    for row in rows:
        idx = int(row["index"])
        if idx in used:
            continue
        try:
            if abs(float(row.get("hours") or 0) - group.total_hours) > 0.01:
                continue
        except Exception:
            continue
        if project_expected and not _project_matches(str(row.get("project") or ""), project_expected):
            continue
        candidates.append(idx)
    if len(candidates) == 1:
        return candidates[0]
    return None


def _reload_effort_page_for_readback(page: Any, url: str) -> None:
    page.goto(url, wait_until="domcontentloaded")
    page.wait_for_timeout(1500)
    try:
        page.wait_for_load_state("networkidle", timeout=8000)
    except Exception:
        pass


def _click_delete_row(page: Any, row_index: int) -> bool:
    clicked = page.evaluate(
        """(rowIndex) => {
          const rows = Array.from(
            document.querySelectorAll('table.table-1 tbody tr, table.colored tbody tr')
          ).filter(tr => tr.querySelectorAll('td').length >= 13);
          const tr = rows[rowIndex];
          if (!tr) return false;
          const cells = Array.from(tr.querySelectorAll('td'));
          const actionCell = cells[cells.length - 1] || tr;
          const candidates = Array.from(actionCell.querySelectorAll('a, button, i, span')).filter(el => {
            const text = (el.innerText || el.textContent || '').trim();
            const title = (el.getAttribute('title') || el.getAttribute('aria-label') || '').trim();
            const cls = (el.className || '').toString();
            const onclick = (el.getAttribute('onclick') || '').toString();
            const href = (el.getAttribute('href') || '').toString();
            const haystack = text + title + cls + onclick + href;
            return text === '×'
              || text === 'x'
              || /删除|移除|取消|叉|delete|remove|del/i.test(haystack)
              || /icon-(remove|trash|delete|close)|btn-(delete|remove)|deleter/i.test(cls);
          });

          const preferred = candidates.find(el => {
            const text = (el.innerText || el.textContent || '').trim();
            const cls = (el.className || '').toString();
            const title = (el.getAttribute('title') || '').toString();
            return text !== '+' && !/plus|add|新增|添加/i.test(cls + title);
          }) || candidates[0];
          if (!preferred) return false;

          const target = preferred.tagName === 'I' || preferred.tagName === 'SPAN'
            ? (preferred.closest('a, button') || preferred)
            : preferred;
          target.click();
          return true;
        }""",
        row_index,
    )
    if clicked:
        page.wait_for_timeout(800)
        for selector in [
            "[class*='modal'] button:has-text('确认')",
            "[class*='dialog'] button:has-text('确认')",
            "[class*='modal'] button:has-text('确定')",
            "[class*='dialog'] button:has-text('确定')",
            ".btn-primary:has-text('确认')",
            ".btn-primary:has-text('确定')",
        ]:
            try:
                loc = page.locator(selector)
                if loc.count() > 0:
                    loc.first.click(timeout=1000)
                    page.wait_for_timeout(800)
                    break
            except Exception:
                continue
    return bool(clicked)


def _is_leave_row(row: dict[str, Any]) -> bool:
    return bool(re.search(r"请假|休假|年假|病假|调休", row.get("all_text", "")))


def _row_has_content(row: dict[str, Any]) -> bool:
    return bool(
        str(row.get("name") or "").strip()
        or float(row.get("hours") or 0) > 0
        or str(row.get("project") or "").strip()
        or str(row.get("description") or "").strip()
    )


def _delete_surplus_rows(page: Any, keep_indexes: set[int]) -> None:
    """删除不属于本次计划的默认/多余工时行。

    禅道会预加载一些默认行。保存前只保留本次填写的行和请假行，
    避免下方多出来的选项一起提交。
    """
    rows = _read_existing_rows(page)
    delete_indexes = [
        int(row["index"])
        for row in rows
        if int(row["index"]) not in keep_indexes
        and not _is_leave_row(row)
    ]
    for row_index in sorted(delete_indexes, reverse=True):
        row_snapshot = next((row for row in rows if int(row["index"]) == row_index), {})
        label = row_snapshot.get("name") or row_snapshot.get("all_text", "")[:30] or f"row {row_index}"
        if _click_delete_row(page, row_index):
            print(f"  🗑️ 已删除多余工时行：{label}")
        else:
            raise RuntimeError(f"需要删除多余工时行但找不到删除控件：{label}")
        rows = _read_existing_rows(page)


def _project_display(choice: ProjectChoice) -> str:
    if choice.source == "fallback":
        return f"{choice.project}（兜底）"
    if choice.project:
        return choice.project
    return f"(下拉搜索：{choice.query})"


def fill_one_day(target_date: Date,
                 entries: list[WorklogEntry],
                 mapping: MappingStore,
                 dry_run: bool = False,
                 save: bool = True) -> None:
    groups = build_zentao_effort_groups(entries, mapping)
    leave_hours = _calc_leave_hours_for_date(entries)

    print(f"\n📅 {target_date} 待填禅道工时")
    if not groups and not leave_hours:
        print("  （无有效工时明细，跳过）")
        return

    for group in groups:
        print(f"  - {group.name}  {group.total_hours:g}h")
        print(f"    项目={_project_display(group.project_choice)}  系统={group.system}  类型={group.task_type}")
        print(f"    描述={group.description}")
    if leave_hours:
        print(f"  - 请假 {leave_hours:g}h（若禅道未自动生成请假行，脚本报错提醒）")

    if dry_run:
        return

    url = ZT_EFFORT_URL.format(date=target_date.isoformat())
    with get_browser_context() as context:
        page = find_or_open_tab(context, ZENTAO_URL_PREFIX, url)
        page.goto(url, wait_until="domcontentloaded")
        page.wait_for_timeout(1800)
        try:
            page.wait_for_load_state("networkidle", timeout=8000)
        except Exception:
            pass

        existing = _read_existing_rows(page)
        if leave_hours > 0:
            leave_rows = [
                row for row in existing
                if re.search(r"请假|休假|年假|病假|调休", row.get("all_text", ""))
            ]
            existing_leave_h = sum(float(row.get("hours") or 0) for row in leave_rows)
            if existing_leave_h + 0.01 < leave_hours:
                if not leave_rows:
                    raise RuntimeError(
                        f"{target_date} 需要请假 {leave_hours:g}h，但禅道没有可识别的请假行/模板。"
                    )
                deficit = leave_hours - existing_leave_h
                row_idx = int(leave_rows[0]["index"])
                existing_h = float(leave_rows[0].get("hours") or 0)
                _set_text_cell(page, row_idx, 7, f"{existing_h + deficit:g}")
                print(f"  ✏️  请假行已更新：{existing_h:g}h -> {existing_h + deficit:g}h")

        used: set[int] = set()
        selected_projects: dict[str, str] = {}
        for group in groups:
            existing = _read_existing_rows(page)
            row_idx = _find_row_index(existing, group.name, used)
            if row_idx is None:
                add_from = max(used) if used else (len(existing) - 1 if existing else None)
                _click_add_row(page, add_from)
                existing = _read_existing_rows(page)
                row_idx = len(existing) - 1
            used.add(row_idx)

            selected_project = _fill_effort_row_on_page(page, row_idx, group)
            selected_projects[group.name] = selected_project
            _validate_filled_row(page, row_idx, group, selected_project)

            if group.project_choice.source == "keyword" and selected_project:
                for entry in group.entries:
                    mapping.learn_task_project(
                        task_name=entry.task_name,
                        project=selected_project,
                        query=group.project_choice.query,
                        system=group.system,
                    )

        _delete_surplus_rows(page, used)

        if save:
            print("  💾 点击保存，提交本页工时确认")
            _save_page_and_confirm(page)
            _reload_effort_page_for_readback(page, url)
            existing_after = _read_existing_rows(page)
            used_after: set[int] = set()
            for group in groups:
                row_idx = _find_saved_row_index(
                    existing_after,
                    group,
                    selected_projects.get(group.name, ""),
                    used_after,
                )
                if row_idx is None:
                    raise RuntimeError(f"保存并刷新后未找到对应行：{group.name}")
                used_after.add(row_idx)
                _validate_saved_row(page, row_idx, group, selected_projects.get(group.name, ""))
            print(f"  ✅ {target_date} 已保存，并通过刷新后回读校验")
        else:
            print(f"  ✅ {target_date} 已填表，未保存（--no-save）")


# ---------------------------------------------------------------------------
# CLI
# ---------------------------------------------------------------------------

def _load_entries(args: argparse.Namespace, start: Date, end: Date) -> list[WorklogEntry]:
    if args.worklog_csv:
        return load_worklog_from_csv(args.worklog_csv, start, end)
    if args.worklog_json:
        return load_worklog_from_json(args.worklog_json, start, end)
    return load_worklog(args.worklog, start, end)


def main() -> None:
    parser = argparse.ArgumentParser(description="根据工作日志填写禅道工时（自包含严格校验）")
    parser.add_argument("--start", required=True, help="YYYY-MM-DD")
    parser.add_argument("--end", required=True, help="YYYY-MM-DD")
    parser.add_argument("--worklog", default=None, help="工作日志 Excel 路径")
    parser.add_argument("--worklog-json", default=None, help="工作日志 JSON 路径；传 '-' 时从 stdin 读取")
    parser.add_argument("--worklog-csv", default=None, help="工作日志 CSV 路径")
    parser.add_argument("--mapping", default=None, help="项目映射 JSON 路径")
    parser.add_argument("--dry-run", action="store_true", help="只打印计划，不打开/填写网页")
    parser.add_argument("--no-save", action="store_true", help="填好后不点击保存")
    args = parser.parse_args()

    start = parse_date_cell(args.start)
    end = parse_date_cell(args.end)
    entries = _load_entries(args, start, end)

    if not entries:
        print(f"❌ {start} ~ {end} 没有读到可填写的工时记录。")
        print("   请确认腾讯文档/企微智能表格已导出，并且至少包含：日期、时间、任务名称。")
        sys.exit(2)

    mapping = MappingStore(args.mapping)
    grouped = group_entries_by_date(entries)

    failures: list[tuple[Date, str]] = []
    succeeded: list[Date] = []
    for day in sorted(grouped):
        try:
            fill_one_day(
                day,
                grouped[day],
                mapping,
                dry_run=args.dry_run,
                save=not args.no_save,
            )
            succeeded.append(day)
        except RuntimeError as exc:
            print(f"\n❌ {day} 填写失败：{exc}")
            print("   跳过该日，继续处理后续日期。")
            failures.append((day, str(exc)))

    if failures:
        print("\n❌ 禅道工时填报部分失败：")
        for day, reason in failures:
            print(f"  - {day}: {reason}")
        sys.exit(2)

    if args.dry_run:
        print(f"\n✅ dry-run 计划生成完成：共 {len(succeeded)} 天。未打开网页，未填写，未保存。")
        return

    if args.no_save:
        print(f"\n✅ 禅道工时已填入页面：共 {len(succeeded)} 天，但按 --no-save 未保存。")
        return

    print(f"\n✅ 禅道工时填报完成：共 {len(succeeded)} 天已填写、保存并回读校验。")


if __name__ == "__main__":
    try:
        main()
    except SystemExit:
        raise
    except Exception as exc:
        print(f"❌ 填写禅道工时失败：{type(exc).__name__}: {exc}")
        sys.exit(1)
